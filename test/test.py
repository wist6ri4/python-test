import openpyxl as px
import random

wb = px.Workbook()
ws = wb.active
sheet = wb.worksheets[0]
ws.title = 'sheet_1'
ws['A1'] = 'Date'
ws['B1'] = 'Product'
ws['C1'] = 'Branch'
ws['D1'] = 'interest'

# B列に製品をランダム生成
for k in range(2, 501):
    a = ('製品1','製品2')
    ws['B' + str(k)] = random.choice(a)

# C列に企業をランダム生成
for m in range(2, 501):
    b = ('支店1', '支店2')
    ws['C' + str(m)] = random.choice(b)

# D列に数字をランダム生成
for n in range(2, 501):
    ws['D' + str(n)] = random.randint(10000, 100000)

# ランダムに生成する月日の範囲を2020-1-1～2020―12-17と決める
from datetime import timedelta
from datetime import date
from random import randrange
import random
start_date = date(2020, 1, 1)
end_date = date(2020, 12, 17)

time_between_dates = end_date - start_date
days_between_dates = time_between_dates.days
random_number_of_days = random.randrange(days_between_dates)
random_date = start_date + timedelta(days=random_number_of_days)

# A列に2020―1-1～2020―12―17の範囲でランダムに生成する
for p in range(2, 501):
    tart_date = date(2020, 1, 1)
    end_date = date(2020, 12, 17)

    time_between_dates = end_date - start_date
    days_between_dates = time_between_dates.days
    random_number_of_days = random.randrange(days_between_dates)
    random_date = start_date + timedelta(days=random_number_of_days)

    ws['A' + str(p)] = (random_date)
    # セルの大きさ調整
    sheet.column_dimensions['A'].width = 13

# 変更内容を保存する
wb.save('sample1.xlsx')